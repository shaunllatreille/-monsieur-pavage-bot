import os, json, logging
from datetime import datetime
from dotenv import load_dotenv
from groq import Groq
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (Application, MessageHandler, CommandHandler,
                          CallbackQueryHandler, filters, ContextTypes)
import openpyxl
from openpyxl.utils import get_column_letter

# Charger env.txt
load_dotenv(dotenv_path="env.txt")

TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
GROQ_API_KEY   = os.getenv("GROQ_API_KEY")
MASTER_FILE    = os.getenv("MASTER_FILE", "Monsieur_Pavage_MASTER_2026.xlsx")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.FileHandler("bot.log", encoding="utf-8"), logging.StreamHandler()]
)
logger = logging.getLogger(__name__)
groq_client = Groq(api_key=GROQ_API_KEY)

CLIENT_KEYWORDS = [
    "client","nom","tel","telephone","adresse","email","courriel",
    "pieds","pi2","sqft","montant","ville","scellant","asphalte","fissure",
    "nettoyage","travaux","installation","patch","entree","c/s","cs"
]

def is_client_msg(text):
    if len(text) < 10: return False
    t = text.lower()
    return sum(1 for kw in CLIENT_KEYWORDS if kw in t) >= 1

def parse_client_info(message, sender):
    prompt = f"""Tu es un assistant CRM pour Monsieur Pavage Quebec.
Extrait les infos du message et retourne UNIQUEMENT un JSON valide sans aucun texte avant ou apres.
Champs: nom, telephone, courriel, adresse, ville, type_service (Scellant C/S|Asphalte chaude|Asphalte froide|Fissures|Scellant + Fissures|Nettoyage|Patch asphalte), pieds_carres (nombre), montant (nombre decimal sans symbole), date_installation (YYYY-MM-DD ou null), priorite (1 - Urgent|2 - Normal|3 - Flexible), notes, categorie (SCELLANT ou ASPHALTE).
Si absent mets null.
Message de {sender}: {message}"""
    try:
        r = groq_client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role":"user","content":prompt}],
            temperature=0.1, max_tokens=600
        )
        text = r.choices[0].message.content.strip().replace("```json","").replace("```","").strip()
        data = json.loads(text)
        return data if sum(1 for v in data.values() if v) >= 2 else None
    except Exception as e:
        logger.error(f"IA error: {e}")
        return None

def save_to_master(data, sender, categorie):
    """Sauvegarde dans le bon onglet du fichier master 2026"""
    try:
        # Charger le fichier master
        master_path = MASTER_FILE
        if not os.path.exists(master_path):
            # Chercher dans Downloads
            master_path = f"C:\\Users\\shaun\\Downloads\\{MASTER_FILE}"
        
        wb = openpyxl.load_workbook(master_path)
        
        # Choisir le bon onglet
        sheet_name = "SCELLANT" if categorie == "SCELLANT" else "ASPHALTE"
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Essayer des variantes du nom
            for name in wb.sheetnames:
                if "SCELLANT" in name.upper() and categorie == "SCELLANT":
                    ws = wb[name]; break
                elif "ASPHALTE" in name.upper() and categorie == "ASPHALTE":
                    ws = wb[name]; break
            else:
                ws = wb.active
        
        # Trouver la prochaine ligne vide
        next_row = ws.max_row + 1
        
        # Calculer les taxes
        montant = float(data.get("montant") or 0)
        tps  = round(montant * 0.05, 2)
        tvq  = round(montant * 0.09975, 2)
        total = round(montant + tps + tvq, 2)
        prefix = "A" if categorie == "ASPHALTE" else "S"
        
        # Écrire les données dans les bonnes colonnes
        # Colonnes: ID, Date entree, Date install, Priorite, Statut, Client, Tel, Courriel, 
        #           Adresse, Ville, Rep, Type service, Pi2, Montant HT, TPS, TVQ, TOTAL,
        #           Paiement, Paye, Solde, Notes, SMS, Courriel conf, Facture, Reviews, Source
        row_data = {
            1:  f"{prefix}-2026-{next_row-1:03d}",
            2:  datetime.now().strftime("%Y-%m-%d"),
            3:  data.get("date_installation") or "",
            4:  data.get("priorite") or "2 - Normal",
            5:  "Nouveau",
            6:  data.get("nom") or "",
            7:  str(data.get("telephone") or ""),
            8:  data.get("courriel") or "",
            9:  data.get("adresse") or "",
            10: data.get("ville") or "",
            11: sender,
            12: data.get("type_service") or "",
            13: data.get("pieds_carres") or "",
            14: montant if montant else "",
            15: tps if montant else "",
            16: tvq if montant else "",
            17: total if montant else "",
            18: "",
            19: "",
            20: total if montant else "",
            21: data.get("notes") or "",
            22: "Non",
            23: "Non",
            24: "Non",
            25: "Non",
            26: "Telegram"
        }
        
        for col, val in row_data.items():
            ws.cell(row=next_row, column=col, value=val)
        
        wb.save(master_path)
        logger.info(f"✅ Master Excel [{sheet_name}] ligne {next_row}: {data.get('nom')}")
        return True, sheet_name
    except Exception as e:
        logger.error(f"❌ Erreur Excel master: {e}")
        return False, ""

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    if not msg or not msg.text: return
    text = msg.text.strip()
    sender = msg.from_user.full_name or msg.from_user.username or "Inconnu"
    if not is_client_msg(text): return
    
    logger.info(f"MSG from {sender}: {text[:60]}")
    data = parse_client_info(text, sender)
    if not data: return
    
    cat = data.get("categorie", "SCELLANT")
    emoji = "🏗️" if cat == "ASPHALTE" else "🖤"
    
    ok, sheet = save_to_master(data, sender, cat)
    
    nom     = data.get("nom") or "—"
    tel     = data.get("telephone") or "—"
    ville   = data.get("ville") or "—"
    svc     = data.get("type_service") or "—"
    pi2     = data.get("pieds_carres") or "—"
    m       = data.get("montant")
    m_str   = f"{float(m):,.2f} $".replace(",", " ") if m else "—"
    date_i  = data.get("date_installation") or "A confirmer"
    saved   = f"📊 {sheet}" if ok else "❌ Erreur"

    txt = (
        f"{emoji} *Client enregistre* | {saved}\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"👤 *Nom:* {nom}\n"
        f"📞 *Tel:* {tel}\n"
        f"📍 *Ville:* {ville}\n"
        f"🏷 *Service:* {svc}\n"
        f"📐 *Pi2:* {pi2}\n"
        f"💰 *Montant:* {m_str}\n"
        f"📅 *Installation:* {date_i}\n"
        f"👨‍💼 *Rep:* {sender}"
    )

    adresse_maps = f"{data.get('adresse','')} {ville}".replace(" ", "+")
    kb = [[
        InlineKeyboardButton("✅ Confirmer", callback_data=f"confirm|{nom}|{tel}"),
        InlineKeyboardButton("🗺️ Route", callback_data=f"route|{adresse_maps}")
    ],[
        InlineKeyboardButton("📄 Facture", callback_data=f"invoice|{nom}|{m or 0}|{svc}"),
        InlineKeyboardButton("❌ Annuler", callback_data=f"cancel|{nom}")
    ]]
    await msg.reply_text(txt.strip(), parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(kb))

async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    parts = q.data.split("|")
    action = parts[0]
    if action == "confirm":
        nom = parts[1] if len(parts)>1 else "client"
        tel = parts[2] if len(parts)>2 else ""
        await q.edit_message_text(f"✅ *{nom}* confirme!\n📞 {tel}\n_Statut mis a jour_", parse_mode="Markdown")
    elif action == "route":
        adr = parts[1] if len(parts)>1 else ""
        url = f"https://www.google.com/maps/dir/?api=1&destination={adr}"
        await q.edit_message_text(f"🗺️ *Route*\n\n[Ouvrir Google Maps]({url})", parse_mode="Markdown")
    elif action == "invoice":
        nom = parts[1] if len(parts)>1 else "Client"
        m   = float(parts[2]) if len(parts)>2 else 0
        svc = parts[3] if len(parts)>3 else ""
        tps = m*0.05; tvq = m*0.09975; tot = m+tps+tvq
        await q.edit_message_text(
            f"📄 *Facture — {nom}*\n━━━━━━━━━━━━\n{svc}\nHT: {m:,.2f} $\nTPS: {tps:,.2f} $\nTVQ: {tvq:,.2f} $\n*TOTAL: {tot:,.2f} $*",
            parse_mode="Markdown")
    elif action == "cancel":
        nom = parts[1] if len(parts)>1 else "client"
        await q.edit_message_text(f"❌ *{nom}* annule.", parse_mode="Markdown")

async def cmd_start(update, context):
    await update.message.reply_text(
        "🤖 *Bot CRM Monsieur Pavage actif!*\n\n"
        "Ecris les infos du client dans le groupe et je m'occupe du reste!\n\n"
        "Commandes:\n/stats - Statistiques\n/aide - Aide",
        parse_mode="Markdown")

async def cmd_stats(update, context):
    try:
        master_path = MASTER_FILE
        if not os.path.exists(master_path):
            master_path = f"C:\\Users\\shaun\\Downloads\\{MASTER_FILE}"
        wb = openpyxl.load_workbook(master_path)
        s = a = 0
        for name in wb.sheetnames:
            ws = wb[name]
            count = sum(1 for row in ws.iter_rows(min_row=3, max_col=6) if row[5].value)
            if "SCELLANT" in name.upper(): s = count
            elif "ASPHALTE" in name.upper(): a = count
        await update.message.reply_text(
            f"📊 *Stats 2026*\n🖤 Scellant: *{s}* clients\n🏗️ Asphalte: *{a}* clients\n👥 Total: *{s+a}*",
            parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"❌ Erreur: {e}")

async def cmd_aide(update, context):
    await update.message.reply_text(
        "📖 *Comment utiliser le bot:*\n\n"
        "Ecris simplement les infos du client, exemple:\n"
        "_Roger Belanger 418-555-1234 14 rue Rouville 800pi2 scellant 450$_\n\n"
        "Le bot detecte automatiquement et enregistre dans le bon onglet!",
        parse_mode="Markdown")

def main():
    if not TELEGRAM_TOKEN:
        logger.error("TELEGRAM_TOKEN manquant dans env.txt"); return
    if not GROQ_API_KEY:
        logger.error("GROQ_API_KEY manquant dans env.txt"); return

    app = Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("stats", cmd_stats))
    app.add_handler(CommandHandler("aide",  cmd_aide))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    logger.info("🚀 Bot CRM Monsieur Pavage demarre!")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
